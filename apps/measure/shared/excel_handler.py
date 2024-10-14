import logging
import enum
import typing as t
from typing import (
    List, Dict,
    Optional,
)

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("mylogger")


class ExcelHandler:
    def __init__(self, filename: str, sheet_name: Optional[str] = None):
        self._filename = filename
        self._sheet_name = sheet_name

        self._workbook: Optional[Workbook] = None
        self._sheet: Optional[Worksheet] = None

        self._min_row: Optional[int] = None

    def get_sheet(self) -> Worksheet:
        return self._sheet

    def get_workbook(self) -> Workbook:
        return self._workbook

    def load_workbook(self):
        """ 加载 Excel 工作簿 """
        try:
            self._workbook = openpyxl.load_workbook(self._filename)
        except FileNotFoundError:
            self._workbook = openpyxl.Workbook()

        if self._sheet_name:
            self._sheet = self._workbook.get_sheet_by_name(self._sheet_name)
        else:
            self._sheet = self._workbook.active

    def save_workbook(self):
        """ 保存工作簿 """
        self._workbook.save(self._filename)

    def close_workbook(self):
        """ 关闭工作簿 """
        self._workbook.close()


class ExcelDataRow:
    def __init__(self, values):
        self._data = values

    def __getitem__(self, item):
        return self._data[item]


class _ExcelHandlerWrapper:
    """ ExcelHandler 的包装类 """

    def __init__(self, base: ExcelHandler):
        # 1. 这个也算是个委托了，但是设置成 _base 私有，将意味着使用本类的客户类不需要知道 _base 的存在。
        # 2. 但是如果委托函数过多，应当将 _base 用 get_base 开放出去，不再使用简单委托函数！
        self._base = base

    def get_base(self):
        return self._base

    def get_sheet(self):
        return self._base.get_sheet()

    def get_workbook(self):
        return self._base.get_workbook()

    def load_workbook(self):
        return self._base.load_workbook()

    def save_workbook(self):
        return self._base.save_workbook()

    def close_workbook(self):
        return self._base.close_workbook()


class ExcelReader(_ExcelHandlerWrapper):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._header = None
        self._reversed_header = None

    def _generator_of_read_data(self, *, min_row=1):
        for row in self.get_sheet().iter_rows(min_row=min_row, values_only=True):
            if not any(row):
                continue
            yield row

    class HeaderRowNumEnum(enum.Enum):
        ONE = 1
        TWO = 2

    def is_valid_excel_header(self, target_headers: List[List[str]]):
        """ 判断是不是有效的模板 """
        for i, target_header in enumerate(target_headers):
            # 这里要注意，cell.value 会为 None
            template_header = ["" if cell.value is None else cell.value for cell in self.get_sheet()[i + 1]]
            if target_header != template_header:
                logger.info("%s", f"\n  target_header: {target_header}\ntemplate_header: {template_header}\n")
                return False
        return True

    def get_table_header(self, header_row_num: HeaderRowNumEnum = HeaderRowNumEnum.TWO) -> Dict[str, int]:
        if self._header:
            return self._header
        if header_row_num == self.HeaderRowNumEnum.ONE:
            raise NotImplementedError("表头的行数为 1 的逻辑尚未未实现")
        # 表头占两行时：第一行有中文，第二行也有中文的时候，第二行才是有用的
        #   2024-09-18：太不舒服了，pandas 库能不能解决这个问题？先用目前这个方式实现吧！
        max_column = self.get_sheet().max_column

        def process_cell(cell):
            if cell.value is None:
                return cell.value
            # 加个判定，目前我认为必须是 str
            if not isinstance(cell.value, str):
                raise TypeError("cell.value 目前只支持字符串类型")
            return cell.value.strip()

        row1 = map(process_cell, self.get_sheet()[1])
        row2 = map(process_cell, self.get_sheet()[2])
        merged_row = ["" for _ in range(max_column)]
        for i, v in enumerate(row1):
            if v:
                merged_row[i] = v
        for i, v in enumerate(row2):
            if v:
                merged_row[i] = v
        res = {v: i for i, v in enumerate(merged_row)}
        self._header = res
        return res

    def get_reversed_table_header(self, header_row_num: HeaderRowNumEnum = HeaderRowNumEnum.TWO) -> Dict[int, str]:
        if self._reversed_header:
            return self._reversed_header
        header = self.get_table_header(header_row_num=header_row_num).items()
        self._reversed_header = {v: k for k, v in header}
        return self._reversed_header

    def read_data(self, *, min_row=1):
        yield from self._generator_of_read_data(min_row=min_row)


class ExcelWriter(_ExcelHandlerWrapper):

    def write_data(self, data: List[List]):
        """ 根据数据添加行内容 """
        for row in data:
            self.get_sheet().append(row)  # 逐行添加数据
