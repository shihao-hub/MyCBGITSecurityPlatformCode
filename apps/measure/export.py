import ast
import collections
import logging
import traceback
from io import BytesIO
import typing as t
from typing import (
    List, Dict,
    Optional, Union, Any, TypeVar, Generic,
    Tuple, Sequence
)
from types import SimpleNamespace

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation

from django.conf import settings

import measure.constants as constant
from measure.models import MissTaskImprove, NewMissTaskDetail
from utils.import_export import ExcelExport

T = TypeVar("T")
logger = logging.getLogger("mylogger")


def fields_mapping(field_val, field_name, reverse=False):
    """字段映射，暂且如此（因为如果像这样写的话，每个分支的相似程度其实很高，不够好）
        reverse：反向映射
    """

    # 2024-09-19：这个函数怎么命名才合适啊？
    # reverse = False，en->cn
    # reverse = True，cn->en

    def throw_none_value_exception(val, msg):
        if val is None:
            raise Exception(msg)

    error_msg = ""

    if field_name == "repeated":
        if not reverse:
            res = "是" if field_val else "否"
        else:
            choices = {"是": True, "否": False}
            res = choices.get(field_val)
            error_msg = f"问题重犯字段必须填 {list(choices.keys())}"
    elif field_name == "confirmed_status":
        if not reverse:
            res = NewMissTaskDetail.CONFIRMED_STATUS_CHOICES.get(field_val)
            error_msg = "超出了确认状态的可选范围"
        else:
            reversed_choices = {v: k for k, v in NewMissTaskDetail.CONFIRMED_STATUS_CHOICES.items()}
            res = reversed_choices.get(field_val)
            error_msg = f"{field_name} 字段必须填 {list(reversed_choices.keys())}"
    else:
        raise ValueError("传入在预期之外的参数")

    if res is None:
        raise ValueError(error_msg)
    return res


class ExcelExporter(Generic[T]):
    """ 超超超类 """
    # 生成字母和列下表对应的字典，形如：0-'A' 1-'B' ... -> chr(index + ord('A')) 不就行了吗？index [0,26)
    col_num_dict = ExcelExport.get_col_num_dict_add()

    def _set_header(self):
        """ 设置表头 """
        raise NotImplementedError

    def _get_field_value_and_style_list(self, task):
        raise NotImplementedError

    def _set_body(self):
        """ 设置正文 """
        raise NotImplementedError

    def _set_footer(self):
        """ 设置其他，此处是类比 enter、do、exit 中的 exit """
        raise NotImplementedError

    def __init__(self,
                 data: List[T],
                 *,
                 sheet_name: Optional[str] = None,
                 sheet_styles: Optional[Dict[str, Any]] = None):

        self._data = data
        self._workbook: Workbook = openpyxl.Workbook()
        self._sheet: Worksheet = self._workbook.active

        if sheet_name:
            self._sheet.title = sheet_name
        if sheet_styles:
            for k, v in sheet_styles.items():
                setattr(self._sheet, k, v)

    def _set_table_row(self, field_value_and_style_list, row, col_num_list):
        sheet = self._sheet
        col_num_list = col_num_list

        for index, value_or_tuple in enumerate(field_value_and_style_list):
            if isinstance(value_or_tuple, tuple):
                value = value_or_tuple[0]
                styles = value_or_tuple[1]
            else:
                value = value_or_tuple
                styles = {}

            column = index + 1
            ExcelExport.set_value_style(
                cell=sheet.cell(row, column),
                value=value,
                **styles
            )
            ExcelExport.record_max_col(col_num_list, value, index)

    def _get_workbook_bytes_io(self):
        sio = BytesIO()
        self._workbook.save(sio)
        sio.seek(0)
        return sio

    def execute(self):
        self._set_header()
        self._set_body()
        self._set_footer()
        return self._get_workbook_bytes_io()


# -------------------------------------------------------------------------------------------------------------------- #
class ExportImprovementsUnderSameTaskSio(ExcelExporter):
    def __init__(self, data, **kwargs):
        super().__init__(data, **kwargs)

        # 2024-10-14：第一次使用我当初重构时搞出来的类，抽象程度太低，存在重复代码，不好用！！
        #            这抽象出来的几个类和函数没多少区别！！！太差了！！！哎...

        self.first_row_header_styles = {
            "font": ExcelExport.header_font,
            "alignment": ExcelExport.alignment_4,
            "fill": ExcelExport.fill_6,  # 黄色的标头填充色
        }

        _first_row_cn_en_mapping = {
            "问题单号": "dts_no",
            "改进措施责任人": "reform_executor",
            "改进类型": "improvement_type",
            "改进措施": "improvement",
            "闭环附件": "attachment_id__id",  # 见 __get_misstaskimprove_res 函数，可以知道原因
            "闭环进展": "close_progress",
            "闭环计划": "close_plan_time",
            "闭环状态": "close_status",
            "是否超期": "is_timeout",
            "更新时间": "update_time",
        }

        self.first_row_header = list(_first_row_cn_en_mapping.keys())

        self.body_start_row = 2
        self.body_row_height = 40

        self.col_num_list = [len(str(x).encode("gb18030")) for x in self.first_row_header]

        if settings.DEBUG:
            self.domain = "https://secevaluation-sit.cbgit.huawei.com/"
        else:
            self.domain = "https://secplatform.cbgit.huawei.com/"

    def _get_field_value_and_style_list(self, task):

        is_timeout = (MissTaskImprove.objects
                      .filter(dts_no=task.dts_no, YN_delete=0)
                      .annotate(**MissTaskImprove.get_condition_of_is_timeout())
                      .values_list("is_timeout", flat=True)
                      .first())
        field_value_and_style_list = (
            task.dts_no[0],  # ["WARN", "WARN", ]
            task.reform_executor,
            task.improvement_type,
            task.improvement,
            (f"{task.file_name}", dict(
                link=f"{self.domain}{task.file_id}?attname={task.file_name}",
                font=ExcelExport.hyperlink_font
            ))
            if task.attachment_id__id
            else "",
            task.close_progress,
            task.close_plan_time,
            task.close_status,
            "是" if is_timeout else "否",
            task.update_time,
        )
        return field_value_and_style_list

    def _set_header(self):
        """ 设置表头 """

        for index, value in enumerate(self.first_row_header):
            row, column = 1, index + 1
            ExcelExport.set_value_style(cell=self._sheet.cell(row, column), value=value,
                                        **self.first_row_header_styles)

    def _set_body(self):
        """ 设置正文 """
        row_delta = 0
        for task in self._data:
            row = self.body_start_row + row_delta
            row_delta += 1

            self._sheet.row_dimensions[row].height = self.body_row_height  # 设置行高

            field_value_and_style_list = self._get_field_value_and_style_list(task)
            self._set_table_row(field_value_and_style_list, row, self.col_num_list)

    def _set_footer(self):
        """ 设置其他，此处是类比 enter、do、exit 中的 exit """
        # 设置自适应列宽
        for i, v in enumerate(self.col_num_list):
            width = min(v, 80)
            self._sheet.column_dimensions[self.col_num_dict.get(i)].width = width + 5


def export_improvements_under_same_task_sio(tasks):
    exporter = ExportImprovementsUnderSameTaskSio(tasks, sheet_styles={
        "font": ExcelExport.font,
        "border": ExcelExport.border,
        "alignment": ExcelExport.alignment_2,
    })
    return exporter.execute()


# -------------------------------------------------------------------------------------------------------------------- #


class SjXwExportMissTaskSio(ExcelExporter):
    """ 送检和现网导出类基类，用于存放二者的共同代码，但是由于继承的复杂性，跳来跳去，确实不好理解 """

    first_row_height = 35
    second_row_height = 40
    body_row_height = 40
    body_start_row = 3
    dts_per_link = f"https://dts-szv.clouddragon.huawei.com/DTSPortal/ticket/"

    def _get_field_value_and_style_list(self, task) -> Sequence[Union[Any, Tuple[Dict]]]:
        """ 获得每行单元格的值和单元格风格 """
        raise NotImplementedError

    def _add_list_box_data_validations(self):
        """ 添加下拉框 """
        raise NotImplementedError

    def _freeze_and_merge(self):
        """ 冻结列 和 合并单元格 """
        raise NotImplementedError

    def __init__(self,
                 first_row_header,
                 second_row_header,
                 data: List[T],
                 *,
                 sheet_name: Optional[str] = None,
                 sheet_styles: Optional[Dict[str, Any]] = None):

        super().__init__(data, sheet_name=sheet_name, sheet_styles=sheet_styles)

        self.first_row_header = first_row_header
        self.second_row_header = second_row_header

        self.col_num_list = [len(str(x).encode("gb18030")) for x in self.first_row_header]  # 记录每列的最大宽度
        self.column_len = len(self.first_row_header)

    @staticmethod
    def _get_body_field_value_by_name(_task, name):
        # 2024-09-19：这个命名也不合适，但是目前我不知道用什么命名好...
        res = ""
        if not getattr(_task, name + "_related") and getattr(_task, name) != "不涉及":
            res = getattr(_task, name)
        elif getattr(_task, name + "_related"):
            res = getattr(_task, name)
        return res

    def _get_sj_xw_common_improvement_fields(self, task):
        this = self
        # 安全架构&方案、安全编码治理、开源治理、特性/用例基线刷新、自动化用例&工具刷新、管理&流程改进
        return {
            "security_architecture": self._get_body_field_value_by_name(task, "security_architecture"),
            "secure_coding": self._get_body_field_value_by_name(task, "secure_coding"),
            "open_source_governance": self._get_body_field_value_by_name(task, "open_source_governance"),
            "case_baseline": self._get_body_field_value_by_name(task, "case_baseline"),
            "auto": self._get_body_field_value_by_name(task, "auto"),
            "manage": self._get_body_field_value_by_name(task, "manage"),
        }

    def _add_list_box_data_validation(self, column: str, formula1):
        list_box_dv = DataValidation(**{
            "type": "list",
            "allow_blank": True,
            "formula1": formula1,
            **constant.COMMON_LIST_BOX_ERROR_INFORMATION
        })
        # 此处 + 2 是因为送检和现网的表头有两行！
        list_box_dv.add(f"{column}{self.body_start_row}:{column}{len(self._data) + 2}")
        self._sheet.add_data_validation(list_box_dv)

    def _set_header(self):
        first_row_header_styles = {
            "font": ExcelExport.header_font,
            "alignment": ExcelExport.alignment_4,
            "fill": ExcelExport.fill_6,  # 黄色的标头填充色
        }
        second_row_header_styles = {
            "font": ExcelExport.header_font,
            "alignment": ExcelExport.alignment_4,
            "fill": ExcelExport.fill_6,  # 黄色的标头填充色
        }

        for index, value in enumerate(self.first_row_header):
            row, column = 1, index + 1
            ExcelExport.set_value_style(cell=self._sheet.cell(row, column), value=value, **first_row_header_styles)

        for index, value in enumerate(self.second_row_header):
            row, column = 2, index + 1
            ExcelExport.set_value_style(cell=self._sheet.cell(row, column), value=value, **second_row_header_styles)

        self._sheet.row_dimensions[1].height = self.first_row_height
        self._sheet.row_dimensions[2].height = self.second_row_height

    def _set_body(self):
        row_delta = 0
        for task in self._data:
            row = self.body_start_row + row_delta
            row_delta += 1

            self._sheet.row_dimensions[row].height = self.body_row_height  # 设置行高

            field_value_and_style_list = self._get_field_value_and_style_list(task)
            self._set_table_row(field_value_and_style_list, row, self.col_num_list)

            # 添加下拉框
            self._add_list_box_data_validations()

    def _set_footer(self):
        # 设置自适应列宽
        for i in range(0, len(self.col_num_list)):
            width = min(self.col_num_list[i], 80)
            self._sheet.column_dimensions[self.col_num_dict.get(i)].width = width + 5
        self._freeze_and_merge()


class SjExportMissTaskSio(SjXwExportMissTaskSio):
    def __init__(self,
                 first_row_header,
                 second_row_header,
                 data: List[T],
                 *,
                 sheet_name: Optional[str] = None,
                 sheet_styles: Optional[Dict[str, Any]] = None):
        super().__init__(first_row_header, second_row_header, data, sheet_name=sheet_name, sheet_styles=sheet_styles)

    def _get_field_value_and_style_list(self, task: NewMissTaskDetail):
        if task.dts_no.startswith('DTS'):
            dts_no_and_styles = (task.dts_no, dict(link=self.dts_per_link + f"{task.dts_no}"))
        else:
            dts_no_and_styles = (task.dts_no, dict())
        improvement_fields = self._get_sj_xw_common_improvement_fields(task)
        product_improve = self._get_body_field_value_by_name(task, "product_improve")
        technical_reason = self._get_body_field_value_by_name(task, "technical_reason")
        process_manage_reason = self._get_body_field_value_by_name(task, "process_manage_reason")

        field_value_and_style_list = (
            dts_no_and_styles,  # 问题单号
            task.description,  # 问题描述
            task.dts_create_time,  # 问题时间
            task.level,  # 级别
            task.problem_type,  # 类型
            task.dts_come,  # 问题来源
            task.version,  # 版本
            task.service,  # 应用
            task.product,  # 部门
            task.control_point,  # 问题发生环节（最佳控制点）
            (technical_reason, dict(alignment=ExcelExport.alignment_1)),  # 技术根因
            (process_manage_reason, dict(alignment=ExcelExport.alignment_1)),  # 管理/流程根因
            (improvement_fields.get("security_architecture"), dict(alignment=ExcelExport.alignment_1)),  # 安全架构&方案
            (improvement_fields.get("secure_coding"), dict(alignment=ExcelExport.alignment_1)),  # 安全编码治理
            (improvement_fields.get("open_source_governance"), dict(alignment=ExcelExport.alignment_1)),  # 开源治理
            (improvement_fields.get("case_baseline"), dict(alignment=ExcelExport.alignment_1)),  # 特性/用例基线刷新
            (improvement_fields.get("auto"), dict(alignment=ExcelExport.alignment_1)),  # 自动化用例&工具刷新
            (improvement_fields.get("manage"), dict(alignment=ExcelExport.alignment_1)),  # 管理&流程改进
            (product_improve, dict(alignment=ExcelExport.alignment_1)),  # 产品改进
            {0: "否", 1: "是"}.get(task.yn_common, 0),  # 是否共性问题
            task.improvement_situation,  # 共性问题排查版本及排查落地情况
            fields_mapping(task.repeated, "repeated"),  # 问题重犯
            fields_mapping(task.confirmed_status, "confirmed_status"),  # 确认状态
            task.analyze_executor,  # 分析责任人
            task.analyze_executor,  # 验收责任人
            {0: "待分析", 1: "分析完成"}.get(task.analyze_status),  # 分析状态
            task.improvement_status,  # 改进状态
            task.acceptance_status,  # 验收状态
        )

        return field_value_and_style_list

    def _add_list_box_data_validations(self):
        self._add_list_box_data_validation("D", constant.SJ_EXCEL_DATA_VALIDATION_FORMULA_OF_LEVEL)
        self._add_list_box_data_validation("E", constant.SJ_EXCEL_DATA_VALIDATION_FORMULA_OF_PROBLEM_TYPE)
        self._add_list_box_data_validation("F", constant.SJ_EXCEL_DATA_VALIDATION_FORMULA_OF_DTS_COME)
        self._add_list_box_data_validation("J", constant.SJ_EXCEL_DATA_VALIDATION_FORMULA_OF_CONTROL_POINT)

    def _freeze_and_merge(self):
        # 冻结首行和首列，主要冻结的是：[问题单号, 最佳控制点]，目前共 10 列
        frozen_max_column = 10

        self._sheet.freeze_panes = "B3"
        for i in range(frozen_max_column):
            char = chr(i + ord('A'))
            self._sheet.merge_cells(range_string=f"{char}1:{char}2")

        # 合并单元格
        merge_cells = [
            "K1:L1",  # 问题根因分析（技术根因，流程/管理根因）
            "M1:S1",  # 改进措施的那些，ord('S') - ord('M') + 1 = 7

            "T1:T2",
            "U1:U2",
            "V1:V2",
            "W1:W2",
            "X1:X2",
            "Y1:Y2",
            "Z1:Z2",
            "AA1:AA2",
            "AB1:AB2",
        ]

        for range_string in merge_cells:
            self._sheet.merge_cells(range_string=range_string)


class XwExportMissTaskSio(SjXwExportMissTaskSio):
    """ 注意，抽象程度还是可以的，只需要实现三个抽象方法即可实现所有功能！ """

    def __init__(self,
                 first_row_header,
                 second_row_header,
                 data: List[T],
                 *,
                 sheet_name: Optional[str] = None,
                 sheet_styles: Optional[Dict[str, Any]] = None):
        super().__init__(first_row_header, second_row_header, data, sheet_name=sheet_name, sheet_styles=sheet_styles)

    def _get_field_value_and_style_list(self, task: NewMissTaskDetail):
        if task.dts_no.startswith('DTS'):
            dts_no_and_styles = (task.dts_no, dict(link=self.dts_per_link + f"{task.dts_no}"))
        else:
            dts_no_and_styles = (task.dts_no, dict())
        improvement_fields = self._get_sj_xw_common_improvement_fields(task)
        security_measures = self._get_body_field_value_by_name(task, "security_measures")
        compliance_policies = self._get_body_field_value_by_name(task, "compliance_policies")
        product_improve = self._get_body_field_value_by_name(task, "product_improve")
        technical_reason = self._get_body_field_value_by_name(task, "technical_reason")
        process_manage_reason = self._get_body_field_value_by_name(task, "process_manage_reason")

        logger.info("%s", f"task.dts_create_time: {task.dts_create_time}")
        field_value_and_style_list = (
            dts_no_and_styles,  # 问题单号
            task.description,  # 问题描述
            task.dts_create_time,  # 问题时间
            task.level,  # 级别
            task.problem_type,  # 类型
            task.dts_come,  # 问题来源
            task.version,  # 版本
            task.service,  # 应用
            task.product,  # 部门
            task.control_point,  # 问题发生环节（最佳控制点）
            (technical_reason, dict(alignment=ExcelExport.alignment_1)),  # 技术根因
            (process_manage_reason, dict(alignment=ExcelExport.alignment_1)),  # 管理/流程根因
            (improvement_fields.get("security_architecture"), dict(alignment=ExcelExport.alignment_1)),  # 安全架构&方案
            (improvement_fields.get("secure_coding"), dict(alignment=ExcelExport.alignment_1)),  # 安全编码治理
            (improvement_fields.get("open_source_governance"), dict(alignment=ExcelExport.alignment_1)),  # 开源治理
            (improvement_fields.get("case_baseline"), dict(alignment=ExcelExport.alignment_1)),  # 特性/用例基线刷新
            (improvement_fields.get("auto"), dict(alignment=ExcelExport.alignment_1)),  # 自动化用例&工具刷新
            (improvement_fields.get("manage"), dict(alignment=ExcelExport.alignment_1)),  # 管理&流程改进
            (security_measures, dict(alignment=ExcelExport.alignment_1)),  # 产品改进
            (compliance_policies, dict(alignment=ExcelExport.alignment_1)),  # 产品改进
            (product_improve, dict(alignment=ExcelExport.alignment_1)),  # 产品改进
            {0: "否", 1: "是"}.get(task.yn_common, 0),  # 是否共性问题
            task.improvement_situation,  # 共性问题排查版本及排查落地情况
            fields_mapping(task.repeated, "repeated"),  # 问题重犯
            fields_mapping(task.confirmed_status, "confirmed_status"),  # 确认状态
            task.analyze_executor,  # 分析责任人
            task.analyze_executor,  # 验收责任人
            {0: "待分析", 1: "分析完成"}.get(task.analyze_status),  # 分析状态
            task.improvement_status,  # 改进状态
            task.acceptance_status,  # 验收状态
        )

        return field_value_and_style_list

    def _add_list_box_data_validations(self):
        self._add_list_box_data_validation("D", constant.XW_EXCEL_DATA_VALIDATION_FORMULA_OF_LEVEL)
        self._add_list_box_data_validation("E", constant.XW_EXCEL_DATA_VALIDATION_FORMULA_OF_PROBLEM_TYPE)
        # #(C)!: self._add_list_box_data_validation("F", constant.XW_EXCEL_DATA_VALIDATION_FORMULA_OF_DTS_COME)
        self._add_list_box_data_validation("J", constant.XW_EXCEL_DATA_VALIDATION_FORMULA_OF_CONTROL_POINT)

    def _freeze_and_merge(self):
        # 冻结首行和首列，主要冻结的是：[问题单号, 最佳控制点]，目前共 10 列
        self._sheet.freeze_panes = "B3"
        for i in range(10):
            char = chr(i + ord('A'))
            self._sheet.merge_cells(range_string=f"{char}1:{char}2")

        # 合并单元格
        merge_cells = [
            "K1:L1",  # 问题根因分析（技术根因，流程/管理根因）
            "M1:U1",  # 改进措施的那些，ord('S') - ord('M') + 1 = 7

            "V1:V2",
            "W1:W2",
            "X1:X2",
            "Y1:Y2",
            "Z1:Z2",
            "AA1:AA2",
            "AB1:AB2",
            "AC1:AC2",
            "AD1:AD2",
        ]

        for range_string in merge_cells:
            self._sheet.merge_cells(range_string=range_string)


def sj_export_misstask_sio(tasks):
    """ 重构方法：函数改名 """
    exporter = SjExportMissTaskSio(constant.SJ_HEADER_OF_FIRST_LINE,
                                   constant.SJ_HEADER_OF_SECOND_LINE,
                                   tasks,
                                   sheet_name=constant.SJ_EXCEL_SHEET_NAME,
                                   sheet_styles={
                                       "font": ExcelExport.font,
                                       "border": ExcelExport.border,
                                       "alignment": ExcelExport.alignment_2,
                                   })
    return exporter.execute()


def xw_export_misstask_sio(tasks):
    """ 重构方法：函数改名 """
    exporter = XwExportMissTaskSio(constant.XW_HEADER_OF_FIRST_LINE,
                                   constant.XW_HEADER_OF_SECOND_LINE,
                                   tasks,
                                   sheet_name=constant.XW_EXCEL_SHEET_NAME,
                                   sheet_styles={
                                       "font": ExcelExport.font,
                                       "border": ExcelExport.border,
                                       "alignment": ExcelExport.alignment_2,
                                   })
    return exporter.execute()
